#!/usr/bin/env python
# -*- coding:utf-8 -*-

import smtplib
import os  
import sys
import datetime  
import time  
import random
import glob  
import shutil  
import string  
import os.path  
import openpyxl
import shutil
from urllib import urlencode
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.MIMEBase import MIMEBase
from email.header import Header
from email import Utils, Encoders

'''
  Send salary mail autoly
'''



#datetime calc
now         = datetime.datetime.now()
curmonth    = int(now.strftime('%m'))  
salarymonth = (curmonth - 1) if (curmonth - 1) > 0 else 12


# input the special month
while True:
    inputMonthStr = raw_input("Enter the month (1 - 12) to send mails; Or 'enter' key for the current month: %d : "%curmonth)
    try:
        inputMonth = int(inputMonthStr)
    except Exception as e:
        #print("Error input:%s"%inputMonthStr, e)
        break
    
    if inputMonth < 1 or inputMonth > 12 :
        continue
    
    inputMonth = (inputMonth - 1) if (inputMonth - 1) > 0 else 12
        
    keys = raw_input("Are you sure the salary month: %d (y/n; Or 'enter' key for OK)? "%inputMonth)

    if len(keys) == 0 or (keys[0] != 'n' and keys[0] != 'N'):
        salarymonth = inputMonth
        break

#excel process
curdir         = os.getcwd().decode('utf-8')
print("Will send mails for the salary month: %d at dir %s ..."%(salarymonth, curdir))

#clear files
salarymonthdir = curdir + os.sep  + "salarys" + os.sep + str(salarymonth)
if os.path.exists(salarymonthdir):
 shutil.rmtree(salarymonthdir)
os.makedirs(salarymonthdir)
salarytemplate_filename = 'template.xlsx'

#find the salary_sum file
xlsxfile_count = 0
xlsx_files     = os.listdir(curdir)
for xlsxfile in xlsx_files:
  if os.path.isfile(xlsxfile) and xlsxfile.find('.xlsx') != -1:
    #print('isFile:' + xlsxfile)
    if xlsxfile.find(salarytemplate_filename) == -1:
      salary_sum_file = xlsxfile
    xlsxfile_count = 1 + xlsxfile_count



if xlsxfile_count > 2:
  print("*.xlsx(s) in the directory exceed two,can't recognize the salary_sum_file!!")
  print('''\
        ****************usage*************************
        comments:please ensure without opened the slary_sum_file when this script is running
        parentdir
        |--any xlsx file(exclude template.xlsx,use as the salary_sum_file)
        |--template.xlsx
        |--smtp.py\
        ''')
  exit(0)

print("meet the request conditions(*.xlsx files:%d),to run...................."%xlsxfile_count)


filename                  = salary_sum_file


#data_only: controls whether cells with formulae have either the formula (default) 
#or the value stored the last time Excel read the sheet   
mybook                    = load_workbook(filename,read_only=True,data_only=True)
sheetnames                = mybook.get_sheet_names()
lastsheetname             = sheetnames[-1]
#default salary month is the last one
salary_sheet_sum          = mybook[lastsheetname]
salary_sheet_sum_cell     = salary_sheet_sum.cell


#mail process
charset                   = 'gb2312'
receiver                  = 'changyunleitest@126.com'
sender                    = 'ai.dangmei@astute-tec.com'
smtpserver                = 'smtp.mxhichina.com'

#sender                   = 'njaixin@163.com'
#smtpserver               = 'smtp.163.com'
username                  = sender

#you must change the password for validity
#password                  = 'programmer@china.com.cn'
password                  = 'programmer@china.com.cn'


send_mail_count           =  0
send_mail_sum             =  0


name_col_pos              = 2
src_finalPayingAmount_pos = 18 #src col
dst_finalPayingAmount_pos = 16 #dst col
src_remark_pos            = src_finalPayingAmount_pos + 2
email_col_pos             = src_finalPayingAmount_pos + 4;

'''
init the copyCells with values: src col --> dst col
copyCells = [
            (2,1),#name
            ...., #copy fields
            (src_col,dst_col),#remarks
            ]
'''
#you should change the value src col --> dst col 
copyCells = [(name_col_pos, 1)] # src name -> dst name 


for dstcol in range(2, dst_finalPayingAmount_pos + 1): 
  copyCells.append((2 + dstcol, dstcol))  # skip the src id col         

copyCells.append((src_remark_pos, dstcol + 1 ))  #src remarks -> dst remarks     

  
#init it to avoid exception when close
smtp         = None          
subject      = u'%d月份工资明细'%salarymonth
content      = u'感谢您的辛勤奉献，您的%d月份工资明细•‿•'%salarymonth

#salary_sheet_sum.max_row is last row, so must add '1' to include the lase item!!!
for row in range(3, salary_sheet_sum.max_row + 1):
  if send_mail_count == 0:
    smtp          = smtplib.SMTP(smtpserver,25)
    smtp.login(username, password) 

  name            = salary_sheet_sum_cell(row=row,column=name_col_pos).value
  
  if not name:
  	print('meet the first empty name cell,reach the end of salary table.......')
  	break



  
  salary_filename        = name + u'.xlsx' 
  
  salary_filename_full   = (salarymonthdir) + (os.sep) 
  salary_filename_full   = salary_filename_full+salary_filename

  #print("filename:" + salary_filename + ",full:" + salary_filename_full)

  shutil.copy(salarytemplate_filename,salary_filename_full)
  salary_book            = load_workbook(salary_filename_full)
  salary_book_sheetnames = salary_book.get_sheet_names()
  salary_book_sheet      = salary_book[salary_book_sheetnames[len(salary_book_sheetnames)-1]]
  
  #copy cell
  for copycelcfg in copyCells:
    
    value                                        = salary_sheet_sum_cell(row=row, column=copycelcfg[0]).value 
    #header_col_name           = salary_sheet_sum_cell(row=2,column=copycelcfg[0]).value
    #dst_header_col_name       =  salary_book_sheet.cell(row=1,column=copycelcfg[1]).value
    #print(u'header name:' + header_col_name +  u',copy value:' + unicode(value) + u',src cell pos:' + unicode(copycelcfg[0]) + u',dst cell pos:' + unicode(copycelcfg[1])  + u',dst head col name:' + dst_header_col_name )
    if not value:
      continue

    salary_book_sheet.cell(row=2,column=copycelcfg[1]).value = value

  #save everyone's salary book  
  salary_book.save(salary_filename_full)  	
  salary_book.close()  
  
  #send email has attached file,must be mixed
  msgRoot            = MIMEMultipart('mixed')
  msgRoot['Subject'] = subject
  msgRoot['From']    = sender   #must field,maybe regard as spam

  #receiver email addr
  email              = salary_sheet_sum_cell(row=row, column=email_col_pos).value
  if not email :
    print('no email address configed')
    continue
  msgRoot['To']      = email

  #content
  msText   = MIMEText(content, _subtype='plain', _charset='utf-8')
  msgRoot.attach(msText)

  #attachment
  maintype = 'application'
  subtype  = 'octet-stream'
  att      = MIMEBase(maintype,subtype)
  att.set_payload(open(salary_filename_full, 'rb').read())
  Encoders.encode_base64(att)
 
  
  #att["Content-Disposition"] = 'attachment; filename= %s'%Header(u'工资明细.xlsx','UTF-8')
  att["Content-Disposition"] = u'attachment; filename=工资明细.xlsx'.encode('gbk')
  
  #print('attachment:' + att["Content-Disposition"])
  msgRoot.attach(att)
  
  send_mail_sum += 1
  print('send mail to:' + name + "; To:" + email + ',No:%d'%send_mail_sum)  
  smtp.sendmail(sender, email, msgRoot.as_string())
  #avoid to be recoginized as spam by email server
  time.sleep(random.randint(1,3))
  send_mail_count = send_mail_count + 1

  #avoid sending too much mails on the same login session will to be recoginized as spam
  if send_mail_count >= 9:
    smtp.quit()
    send_mail_count = 0
    smtp            = None
    print('sleep,will begin the next send mail loop.Please wait...')
    time.sleep(random.randint(12,60))

#finish operation
if smtp :
  smtp.quit()

mybook.close()
print('send salary emails ends!')








